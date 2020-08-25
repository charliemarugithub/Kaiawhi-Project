import tkinter as tk

import openpyxl as xl

from collections import defaultdict, Counter

from tkinter import ttk, filedialog, messagebox

from openpyxl.styles.borders import BORDER_THICK

from openpyxl.utils.exceptions import InvalidFileException

from functions import file_not_found, no_filename, no_sheet_name

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from functions import no_destination_file, packing_report_generated, delivery_report_generated

# creating instance of TK class
root = tk.Tk()
# creating main form window and packing it
main_form = tk.Canvas(root, width=500, height=500)
main_form.pack()
main_form.configure(background="#7289f2")
# name root title page
root.title('Kaiawhi Program')

# creating label for filename and placing it in root
filename_label = tk.Label(root, text='Select Source File Name: ', bg='#7289f2', font="Helvetica 16")
main_form.create_window(250, 40, window=filename_label)

filename = ''

# creating entry label for filename and placing it in root
filename_entry = tk.Entry(root, font="Helvetica, 16")
main_form.create_window(250, 80, window=filename_entry, width=380, height=25)

# creating label for sheet name and placing it in root
sheet_name_label = tk.Label(root, text='Enter New Sheet Name:', bg='#7289f2', font="Helvetica 16")
main_form.create_window(250, 120, window=sheet_name_label)

# creating entry for sheet name  and placing it in root
sheet_name_example_label = tk.Label(root, text='eg 10 Aug Packing List', bg='#7289f2', font="Helvetica 16")
main_form.create_window(250, 150, window=sheet_name_example_label)

# creating entry for sheet name and placing it in root
sheet_name_entry = tk.Entry(root, font="Helvetica, 16")
main_form.create_window(250, 190, window=sheet_name_entry, width=380, height=25)

# creating entry for destination file
destination_label = tk.Label(root, text='Select Or Enter Destination File Name: ', bg='#7289f2', font="Helvetica 16")
main_form.create_window(250, 240, window=destination_label)

# creating entry for filename destination $ placing it in root
destination_entry = tk.Entry(root, font="Helvetica, 16")
main_form.create_window(250, 285, window=destination_entry, width=380, height=25)
# destination_entry.config(state=tk.DISABLED)


# creating menu bar
menubar = tk.Menu(root)
root.config(menu=menubar)

# Create the submenu

subMenu = tk.Menu(menubar, tearoff=0)


# browse computer for source file
def browse_file():
    global filename
    filename = filedialog.askopenfilename(initialdir='c:\\', title='Open File',
                                          filetypes=(('Excel Files', '*.xlsx'), ('All Files', '*.*')))
    filename_entry_label = filename
    print('File path is ' + filename_entry_label)


'''
# Save As File
def save_as_file():
    save_file = filedialog.save(defaultextension='*.*', initialdir='c:\\', title='Save File', filetypes=(('Excel Files', '*.xlsx'), ('All Files', '*.*')))
    print(save_file)
'''


def destination_file():
    global dest_filename
    dest_filename = filedialog.askopenfilename()
    destination_entry['text'] = dest_filename
    print(dest_filename)


menubar.add_cascade(label="File", menu=subMenu)
subMenu.add_command(label="Open  Source File", command=browse_file)
subMenu.add_command(label="Save As", command=destination_file)
subMenu.add_command(label="Exit Program", command=root.destroy)


def about_app():
    tk.messagebox.showinfo('About Kaiawhi App', 'Build Version 1.0 23 Aug 2020: \n'
                                                'Automation Program specific to Kaiawhi Packing and Delivery Reports.')


subMenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label="Help", menu=subMenu)
subMenu.add_command(label="About Us", command=about_app)


def make_packing_list():
    # get method for filename entry
    get_file = filename_entry.get()
    get_destination = destination_entry.get()
    '''
    # check if destination file exists
    if not os.path.isfile(f'{get_destination.strip()}.xlsx'):
        no_destination_file()
    '''

    try:
        # loading workbook on local computer c drive using filename
        wb = xl.load_workbook(f'{get_file}')

        # working with sheet1 on wb 'workbook'
        sheet = wb['Form responses 3']

        # get method for new sheet name
        sheet_name = sheet_name_entry.get()
        new_sheet = wb.create_sheet("Sheet A", 0)
        new_sheet.title = sheet_name
        sheet_name = wb.active

        # deleting columns so that columns required are left for new file
        sheet.delete_cols(1, 12)
        sheet.delete_cols(2)
        sheet.insert_cols(7)
        sheet.delete_cols(2)

        # moving Box Order to column 6 for better reading of numbers
        for cell in sheet['C:C']:
            sheet.cell(row=cell.row, column=6, value=cell.value)

        # deleting old Box Order suburb columnnow as not required
        sheet.delete_cols(3)
        # deleting end columns as not required now
        sheet.delete_cols(9, 6)

        # updating Column Names
        sheet['B1'].value = "Total"
        sheet['C1'].value = "Children"
        sheet['D1'].value = "Adults"
        sheet['F1'].value = "Packing Instructions"
        sheet['G1'].value = "Are there any items you dont want included?"

        # calculate total number of rows and columns in source excel file
        max_rows = sheet.max_row
        max_columns = sheet.max_column

        # setting font for row 1
        bold_font = Font(name='Arial', size=14, bold=True)
        # setting font for text in whole sheet except top row
        cell_font = Font(name='Calibri', size=16)
        # cell alignment to center
        horizon_center = Alignment(horizontal='center')
        # wrap text alignment
        wrap_text = Alignment(wrap_text=True)
        # setting border types
        border = Border(
            left=Side(border_style=BORDER_THICK, color='a8a1ad'),
            right=Side(border_style=BORDER_THICK, color='a8a1ad'),
            top=Side(border_style=BORDER_THICK, color='a8a1ad'),
            bottom=Side(border_style=BORDER_THICK, color='a8a1ad')
        )
        # setting colors for each suburb
        col_panmure = PatternFill(fgColor='80e098', fill_type='solid')
        col_ptengland = PatternFill(fgColor='d9b36c', fill_type='solid')
        col_gi = PatternFill(fgColor='8d9cf0', fill_type='solid')
        col_ruapotaka = PatternFill(fgColor='77d1c8', fill_type='solid')
        col_gifc = PatternFill(fgColor='87d0ed', fill_type='solid')
        col_tamaki = PatternFill(fgColor='bebdf2', fill_type='solid')

        # copying the cell values from source excel file to destination excel file
        for i in range(1, max_rows):
            for j in range(1, max_columns):
                # reading cell value from source excel file
                c = sheet.cell(row=i, column=j)

                # writing the read value to destination excel file
                sheet_name.cell(row=i, column=j).value = c.value
                sheet_name.cell(row=i, column=j).font = cell_font

                # for loop to iterate over Suburbs columns to determine color
                for row in sheet_name.iter_rows(max_row=max_rows, max_col=1):
                    for cell in row:
                        if cell.value == 'Glen Innes':
                            sheet_name.cell(row=i, column=j).fill = col_gi
                        if cell.value == 'Panmure':
                            sheet_name.cell(row=i, column=j).fill = col_panmure
                        if cell.value == 'Point England':
                            sheet_name.cell(row=i, column=j).fill = col_ptengland
                        if cell.value == 'Ruapotaka Marae':
                            sheet_name.cell(row=i, column=j).fill = col_ruapotaka
                        if cell.value == 'GIFC':
                            sheet_name.cell(row=i, column=j).fill = col_gifc
                        if cell.value == 'Tamaki College':
                            sheet_name.cell(row=i, column=j).fill = col_tamaki

                # making row 1 bold font
                sheet_name.cell(row=1, column=i).font = bold_font
                # text alignment for all rows
                sheet_name.cell(row=i, column=j).alignment = horizon_center
                # setting all row height to 45
                sheet_name.row_dimensions[i].height = 45
                # setting borders for all cells
                sheet_name.cell(row=i, column=j).border = border
                # wrapping text on columns 5-8
                sheet_name.cell(row=i, column=5).alignment = wrap_text
                sheet_name.cell(row=i, column=6).alignment = wrap_text
                sheet_name.cell(row=i, column=7).alignment = wrap_text
                sheet_name.cell(row=i, column=8).alignment = wrap_text
                # Boxes column made bold and center
                sheet_name.cell(row=i, column=5).font = bold_font
                sheet_name.cell(row=i, column=5).alignment = horizon_center
                # setting specific column widths
                sheet_name.column_dimensions['A'].width = 20
                sheet_name.column_dimensions['B'].width = 12
                sheet_name.column_dimensions['C'].width = 12
                sheet_name.column_dimensions['D'].width = 12
                sheet_name.column_dimensions['E'].width = 16
                sheet_name.column_dimensions['F'].width = 45
                sheet_name.column_dimensions['G'].width = 45
                sheet_name.column_dimensions['H'].width = 45

                for total in sheet_name.values:
                    for values in total:
                        print(values)
                        '''
                        if total.value == 1 and total.value <= 6:
                            box.value = 1
                        elif total.value == 7 and total.value <= 10:
                            box.value = 2
                        elif total.value > 10:
                            box.value = 3
                        else:
                            box.value = 0
                        '''
        wb.remove(sheet)

        # saving new worksheet to desktop with name packing_list
        wb.save(f'{get_destination}.xlsx')
        packing_button.config(state=tk.DISABLED)
        sheet_name_entry.delete(0, tk.END)
        # destination_entry.delete(9, tk.END)
        packing_report_generated()

    except FileNotFoundError:
        file_not_found()

    except InvalidFileException:
        no_filename()

    except ValueError:
        no_sheet_name()


def make_delivery_list():
    get_file = filename_entry.get()
    get_destination = destination_entry.get()
    '''
    # check if destination field is empty
    if get_destination == '':
        no_destination_file()
    else:
    '''
    try:
        # loading workbook on local computer c drive using filename
        wb = xl.load_workbook(f'{get_file}')

        # working with sheet1 on wb 'workbook'
        sheet = wb['Form responses 3']

        # get method for new sheet name
        sheet_name = sheet_name_entry.get()
        new_sheet = wb.create_sheet("Sheet A", 0)
        new_sheet.title = sheet_name
        sheet_name = wb.active

        # deleting columns so that columns required are left for new file
        sheet.delete_cols(1, 7)
        sheet.delete_cols(9, 4)
        sheet.insert_cols(1)
        sheet.delete_cols(3)
        sheet.delete_cols(9, 1)

        # moving suburbs to column 1. This is required to make suburb color work
        for cell in sheet['F:F']:
            sheet.cell(row=cell.row, column=1, value=cell.value)
        # deleting old suburbs column as not required now
        sheet.delete_cols(6)
        sheet.delete_cols(8, 8)
        sheet.insert_cols(8)
        # updating Column Names
        sheet['B1'].value = "First Name"
        sheet['G1'].value = "Delivery Instructions"
        sheet['H1'].value = "Box Order"

        # calculate total number of rows and columns in source excel file
        max_rows = sheet.max_row
        max_columns = sheet.max_column

        # setting variables for loop
        bold_font = Font(name='Arial', size=14, bold=True)
        # setting font for text in whole sheet except top row
        cell_font = Font(name='Calibri', size=16)
        # cell alignment to center
        horizon_center = Alignment(horizontal='center')
        # wrap text alignment
        wrap_text = Alignment(wrap_text=True)
        border = Border(
            left=Side(border_style=BORDER_THICK, color='a8a1ad'),
            right=Side(border_style=BORDER_THICK, color='a8a1ad'),
            top=Side(border_style=BORDER_THICK, color='a8a1ad'),
            bottom=Side(border_style=BORDER_THICK, color='a8a1ad')
        )
        # setting colors for each suburb
        col_panmure = PatternFill(fgColor='80e098', fill_type='solid')
        col_ptengland = PatternFill(fgColor='d9b36c', fill_type='solid')
        col_gi = PatternFill(fgColor='8d9cf0', fill_type='solid')
        col_ruapotaka = PatternFill(fgColor='77d1c8', fill_type='solid')
        col_gifc = PatternFill(fgColor='87d0ed', fill_type='solid')
        col_tamaki = PatternFill(fgColor='bebdf2', fill_type='solid')

        # copying the cell values from source excel file to destination excel file
        for i in range(1, max_rows):
            for j in range(1, max_columns):
                # reading cell value from source excel file
                c = sheet.cell(row=i, column=j)

                # writing the read value to destination excel file
                sheet_name.cell(row=i, column=j).value = c.value
                sheet_name.cell(row=i, column=j).font = cell_font

                # for loop to iterate over Suburbs columns to determine color
                for row in sheet_name.iter_rows(max_row=max_rows, max_col=max_columns):
                    for cell in row:
                        if cell.value == 'Glen Innes':
                            sheet_name.cell(row=i, column=j).fill = col_gi
                        if cell.value == 'Panmure':
                            sheet_name.cell(row=i, column=j).fill = col_panmure
                        if cell.value == 'Point England':
                            sheet_name.cell(row=i, column=j).fill = col_ptengland
                        if cell.value == 'Ruapotaka Marae':
                            sheet_name.cell(row=i, column=j).fill = col_ruapotaka
                        if cell.value == 'GIFC':
                            sheet_name.cell(row=i, column=j).fill = col_gifc
                        if cell.value == 'Tamaki College':
                            sheet_name.cell(row=i, column=j).fill = col_tamaki

                # making row 1 bold font
                sheet_name.cell(row=1, column=i).font = bold_font
                # wrapping text on columns 3 & 7
                sheet_name.cell(row=i, column=3).alignment = wrap_text
                sheet_name.cell(row=i, column=7).alignment = wrap_text
                # text alignment for all rows
                sheet_name.cell(row=i, column=j).alignment = horizon_center
                sheet_name.cell(row=1, column=j).alignment = horizon_center
                # setting all row height to 30
                sheet_name.row_dimensions[i].height = 65
                # setting borders for all cells
                sheet_name.cell(row=i, column=j).border = border
                # Totals column made bold
                sheet_name.cell(row=i, column=8).font = bold_font
                # setting specific column widths
                sheet_name.column_dimensions['A'].width = 18
                sheet_name.column_dimensions['B'].width = 18
                sheet_name.column_dimensions['C'].width = 24
                sheet_name.column_dimensions['D'].width = 30
                sheet_name.column_dimensions['E'].width = 25
                sheet_name.column_dimensions['F'].width = 25
                sheet_name.column_dimensions['G'].width = 45
                sheet_name.column_dimensions['H'].width = 15
                sheet_name.column_dimensions['I'].width = 45

        # saving new worksheet to desktop with name packing_list
        wb.remove(sheet)
        wb.save(f'{get_destination}.xlsx')
        delivery_button.config(state=tk.DISABLED)
        sheet_name_entry.delete(0, tk.END)
        destination_entry.delete(0, tk.END)
        delivery_report_generated()

    except FileNotFoundError:
        file_not_found()

    except InvalidFileException:
        no_filename()

    except ValueError:
        no_sheet_name()


def clear_all():
    delivery_button.config(state=tk.ACTIVE)
    packing_button.config(state=tk.ACTIVE)
    sheet_name_entry.delete(0, tk.END)
    filename_entry.delete(0, tk.END)
    destination_entry.delete(0, tk.END)


packing_button = ttk.Button(text='Packing List', command=make_packing_list)
main_form.create_window(150, 360, window=packing_button, height=50, width=150)

delivery_button = ttk.Button(text='Delivery List', command=make_delivery_list)
main_form.create_window(350, 360, window=delivery_button, height=50, width=150)

clear_button = ttk.Button(text='CLEAR ALL', command=clear_all)
main_form.create_window(250, 450, window=clear_button, height=50, width=150)

root.mainloop()
