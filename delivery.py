import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.borders import BORDER_THICK

# creating a variable to take any filename input from user
filename = input('Enter Delivery List filename here: ')
new_sheet_name = input('Enter new sheet name for Delivery List: ')

# loading workbook on local computer c drive using filename
wb = xl.load_workbook(f'c:\\Users\\Charlie\\Desktop\\{filename}.xlsx')

# working with sheet1 on wb 'workbook'
sheet = wb['Form responses 3']

new_sheet = wb.create_sheet("Sheet A", 0)
new_sheet.title = new_sheet_name
new_sheet_name = wb.active


# defining delivery list function
def delivery_list():
    # deleting columns so that columns required are left for new file
    sheet.delete_cols(1, 7)
    sheet.delete_cols(9, 4)

    # updating Column Names
    sheet['A1'].value = "First Name"
    sheet['C1'].value = "Address Number"
    sheet['G1'].value = "Delivery Instructions"
    sheet['H1'].value = "Total"

    # calculate total number of rows and columns in source excel file
    max_rows = sheet.max_row
    max_columns = sheet.max_column

    # setting variables for loop
    bold_font = Font(name='Arial', size=12, bold=True)
    # cell alignment to center
    horizon_center = Alignment(horizontal='center')
    # wrap text alignment
    wrap_text = Alignment(wrap_text=True)
    border = Border(
                left=Side(border_style=BORDER_THICK, color='a8a1ad'),
                right=Side(border_style=BORDER_THICK, color='a8a1ad'),
                top=Side(border_style=BORDER_THICK, color='a8a1ad'),
                bottom=Side(border_style=BORDER_THICK, color='a8a1ad')
                )
    # setting colors for each suburb
    col_panmure = PatternFill(fgColor='80e098', fill_type='solid')
    col_ptengland = PatternFill(fgColor='d9b36c', fill_type='solid')
    col_gi = PatternFill(fgColor='8d9cf0', fill_type='solid')
    col_stjohns = PatternFill(fgColor='ba6cd9', fill_type='solid')
    col_glendowie = PatternFill(fgColor='d98aed', fill_type='solid')
    col_mtwell = PatternFill(fgColor='a1f0a9', fill_type='solid')
    col_greenlane = PatternFill(fgColor='f0daa1', fill_type='solid')
    col_mangere = PatternFill(fgColor='e4f0a1', fill_type='solid')
    col_pakuranga = PatternFill(fgColor='e4e86d', fill_type='solid')
    # setting color for 'Total' column
    totals_color = PatternFill(fgColor='f23b38', fill_type='solid')

    # copying the cell values from source excel file to destination excel file
    for i in range(1, max_rows + 1):
        for j in range(1, max_columns + 1):
            # reading cell value from source excel file
            c = sheet.cell(row=i, column=j)

            # writing the read value to destination excel file
            new_sheet.cell(row=i, column=j).value = c.value

            for row in new_sheet_name['A1:J100']:
                for cell in row:
                    if cell.value == 'Panmure':
                        cell.fill = col_panmure
                    if cell.value == 'Pt England':
                        cell.fill = col_ptengland
                    if cell.value == 'Point England':
                        cell.fill = col_ptengland
                    if cell.value == 'Glen Innes':
                        cell.fill = col_gi
                    if cell.value == 'St Johns':
                        cell.fill = col_stjohns
                    if cell.value == 'Glendowie':
                        cell.fill = col_glendowie
                    if cell.value == 'Mt Wellington':
                        cell.fill = col_mtwell
                    if cell.value == 'Greenlane':
                        cell.fill = col_greenlane
                    if cell.value == 'Mangere':
                        cell.fill = col_mangere
                    if cell.value == 'Pakuranga':
                        cell.fill = col_pakuranga

            # making row 1 bold font
            new_sheet.cell(row=1, column=i).font = bold_font
            # text alignment for all rows
            new_sheet.cell(row=i, column=j).alignment = horizon_center
            # setting all row height to 30
            new_sheet_name.row_dimensions[i].height = 30
            # setting borders for all cells
            new_sheet.cell(row=i, column=j).border = border
            # wrapping text on columns 9
            new_sheet.cell(row=i, column=9).alignment = wrap_text
            # setting 'Totals' color column to red and bold font
            new_sheet.cell(row=i, column=8).fill = totals_color
            new_sheet.cell(row=i, column=8).font = bold_font
            # setting specific column widths
            new_sheet_name.column_dimensions['A'].width = 18
            new_sheet_name.column_dimensions['B'].width = 28
            new_sheet_name.column_dimensions['C'].width = 35
            new_sheet_name.column_dimensions['D'].width = 35
            new_sheet_name.column_dimensions['E'].width = 30
            new_sheet_name.column_dimensions['F'].width = 25
            new_sheet_name.column_dimensions['G'].width = 60
            new_sheet_name.column_dimensions['H'].width = 12
            new_sheet_name.column_dimensions['I'].width = 75

    # saving new worksheet to desktop with name packing_list
    wb.remove_sheet(sheet)
    wb.save('c:\\Users\\Charlie\\Desktop\\delivery_list.xlsx')
