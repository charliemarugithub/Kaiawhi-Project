import openpyxl as xl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

# creating a variable to take any filename input from user
filename = input('Enter Packing List filename here: ')
new_sheet_name = input('Enter new sheet name for Packing List: ')

# loading workbook on local computer c drive using filename
wb = xl.load_workbook(f'c:\\Users\\Charlie\\Desktop\\{filename}.xlsx')

# working with sheet1 on wb 'workbook'
sheet = wb['Form responses 3']

new_sheet = wb.create_sheet("Sheet A", 0)
new_sheet.title = new_sheet_name
new_sheet_name = wb.active


# defining packing list function
def packing_list():
    # deleting columns so that columns required are left for new file
    sheet.delete_cols(1, 9)
    sheet.delete_cols(5)

    # updating Column Names
    sheet['E1'].value = "Total"
    sheet['F1'].value = "Children"
    sheet['G1'].value = "Adults"
    sheet['I1'].value = "Are there any items you dont want included?"

    # calculate total number of rows and columns in source excel file
    max_rows = sheet.max_row
    max_columns = sheet.max_column

    # setting variables for loop
    bold_font = Font(name='Arial', size=12, bold=True)
    # cell alignment to center
    horizon_center = Alignment(horizontal='center')
    # row 1 align cells to vertical center
    vertical_center = Alignment(vertical='center')
    # wrap text alignment
    wrap_text = Alignment(wrap_text=True)
    # setting colors for each suburb
    col_panmure = PatternFill(fgColor='80e098', fill_type='solid')
    col_ptengland = PatternFill(fgColor='d9b36c', fill_type='solid')
    col_gi = PatternFill(fgColor='8d9cf0', fill_type='solid')
    col_stjohns = PatternFill(fgColor='ba6cd9', fill_type='solid')

    # copying the cell values from source excel file to destination excel file
    for i in range(1, max_rows + 1):
        for j in range(1, max_columns + 1):
            # reading cell value from source excel file
            c = sheet.cell(row=i, column=j)

            # writing the read value to destination excel file
            new_sheet.cell(row=i, column=j).value = c.value

            for row in new_sheet_name['A1:J100']:
                for cell in row:
                    if cell.value == 'Panmure':
                        cell.fill = col_panmure
                    if cell.value == 'Pt England':
                        cell.fill = col_ptengland
                    if cell.value == 'Point England':
                        cell.fill = col_ptengland
                    if cell.value == 'Glen Innes':
                        cell.fill = col_gi
                    if cell.value == 'St Johns':
                        cell.fill = col_stjohns

            # making row 1 bold font
            new_sheet.cell(row=1, column=i).font = bold_font
            # text alignment for all rows
            new_sheet.cell(row=i, column=j).alignment = horizon_center
            # setting all row height to 30
            new_sheet_name.row_dimensions[i].height = 30
            # wrapping text on columns 8-10
            new_sheet.cell(row=i, column=8).alignment = wrap_text
            new_sheet.cell(row=i, column=9).alignment = wrap_text
            new_sheet.cell(row=i, column=10).alignment = wrap_text
            # setting specific column widths
            new_sheet_name.column_dimensions['A'].width = 21.5
            new_sheet_name.column_dimensions['B'].width = 35
            new_sheet_name.column_dimensions['C'].width = 27
            new_sheet_name.column_dimensions['D'].width = 25
            new_sheet_name.column_dimensions['E'].width = 9
            new_sheet_name.column_dimensions['F'].width = 12.5
            new_sheet_name.column_dimensions['G'].width = 9.8
            new_sheet_name.column_dimensions['H'].width = 75
            new_sheet_name.column_dimensions['I'].width = 75
            new_sheet_name.column_dimensions['J'].width = 75

    # saving new worksheet to desktop with name packing_list
    wb.remove_sheet(sheet)
    wb.save('c:\\Users\\Charlie\\Desktop\\packing_list.xlsx')
